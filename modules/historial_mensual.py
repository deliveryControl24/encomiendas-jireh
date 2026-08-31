"""
==============================================
  modules/historial_mensual.py
  Resumen historico de pedidos por mes
  con cards, exportar Excel y detalle
==============================================
"""

import tkinter as tk
from tkinter import ttk, messagebox
from modules.config import PAGINA_TAMANO
import os
import sys

try:
    from modules.config import get_base_dir
except ImportError:
    def get_base_dir():
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


class HistorialMensualFrame(tk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent, bg="#f5f5f0")
        self.db = db
        self.app = app
        self._pagina = 1
        self._por_pagina = PAGINA_TAMANO
        self._total_paginas = 1
        self._mes_sel = None
        self._datos_cache = []
        self._build()

    def _build(self):
        c = self.app.colores

        hdr = tk.Frame(self, bg=c.get("card_bg", "#ffffff"), pady=14, padx=20)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📅  Historial mensual de pedidos",
                 font=("Segoe UI", 14, "bold"),
                 bg=c.get("card_bg", "#ffffff"),
                 fg=c.get("accent", "#0f6e56")).pack(side="left")

        tk.Button(hdr, text="📥  Exportar Excel",
                  font=("Segoe UI", 9, "bold"),
                  bg=c.get("accent", "#0f6e56"), fg="#ffffff",
                  bd=0, padx=14, pady=6, cursor="hand2",
                  command=self._exportar_excel).pack(side="right", padx=(8, 0))

        tk.Button(hdr, text="🖨  Imprimir",
                  font=("Segoe UI", 9),
                  bg="#f0efe8", fg="#5f5e5a",
                  bd=0, padx=14, pady=6, cursor="hand2",
                  command=self._imprimir).pack(side="right")

        # ── Cards resumen ────────────────────────────────────────────────────
        self.cards_frame = tk.Frame(self, bg=c.get("bg", "#f5f5f0"))
        self.cards_frame.pack(fill="x", padx=14, pady=(10, 0))

        self._cards = {}
        labels_cards = [
            ("meses",      "📅  Meses",        "0",     "#0f6e56", "#e1f5ee"),
            ("envios",     "📦  Total envíos",  "0",     "#0c447c", "#e6f1fb"),
            ("vendido",    "💰  Total vendido",  "$0.00", "#7a4800", "#faeeda"),
            ("pagado",     "✅  Pagado",         "$0.00", "#085041", "#f6fdf9"),
            ("pendiente",  "⏳  Pendiente",      "$0.00", "#791f1f", "#fcebeb"),
        ]
        for key, titulo, valor, fg_color, bg_color in labels_cards:
            card = tk.Frame(self.cards_frame, bg=bg_color,
                            highlightthickness=1, highlightbackground=fg_color + "33",
                            padx=14, pady=10)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8))
            tk.Label(card, text=titulo, font=("Segoe UI", 9),
                     bg=bg_color, fg=fg_color).pack(anchor="w")
            lbl_valor = tk.Label(card, text=valor,
                                 font=("Segoe UI", 14, "bold"),
                                 bg=bg_color, fg=fg_color)
            lbl_valor.pack(anchor="w", pady=(2, 0))
            self._cards[key] = lbl_valor

        # ── Pane principal ───────────────────────────────────────────────────
        paned = tk.PanedWindow(self, orient="horizontal",
                                bg="#d8d6cf", sashwidth=5, sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=14, pady=(8, 12))

        left = tk.Frame(paned, bg="#ffffff",
                        highlightthickness=1, highlightbackground="#dddbd4")
        paned.add(left, minsize=380)

        self._build_tree_resumen(left)
        self._build_paginacion(left)

        self._detalle_frame = tk.Frame(paned, bg="#ffffff",
                                        highlightthickness=1, highlightbackground="#dddbd4")
        paned.add(self._detalle_frame, minsize=300)
        self._build_detalle_vacio()

        self.refresh()

    def _build_tree_resumen(self, parent):
        style = ttk.Style()
        style.configure("Mes.Treeview",
                         font=("Segoe UI", 9), rowheight=30,
                         background="#ffffff", fieldbackground="#ffffff",
                         borderwidth=0)
        style.configure("Mes.Treeview.Heading",
                         font=("Segoe UI", 9, "bold"),
                         background="#0f6e56", foreground="#ffffff",
                         relief="flat")
        style.map("Mes.Treeview",
                  background=[("selected", "#d4f5eb")],
                  foreground=[("selected", "#085041")])

        cols = ("mes", "anio", "envios", "total_vendido", "total_pagado", "pendiente")
        self.tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse",
                                  style="Mes.Treeview")
        hdrs = [
            ("mes",           "Mes",          100, "w"),
            ("anio",          "Año",           55, "center"),
            ("envios",        "Envíos",        60, "center"),
            ("total_vendido", "Vendido",       90, "e"),
            ("total_pagado",  "Pagado",        90, "e"),
            ("pendiente",     "Pendiente",     90, "e"),
        ]
        for col, txt, w, anc in hdrs:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, minwidth=40, anchor=anc)

        self.tree.tag_configure("par",      background="#f9f9f6")
        self.tree.tag_configure("impar",    background="#ffffff")
        self.tree.tag_configure("con_deuda", foreground="#791f1f")

        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_seleccionar)

    def _build_paginacion(self, parent):
        c = self.app.colores
        pag = tk.Frame(parent, bg=c.get("card_bg", "#f0efe8"), pady=5)
        pag.pack(fill="x")
        self._pag_label = tk.Label(pag, text="", font=("Segoe UI", 8),
                                    bg=c.get("card_bg", "#f0efe8"), fg="#888780")
        self._pag_label.pack(side="left", padx=10)

        def _pbtn(txt, cmd):
            return tk.Button(pag, text=txt, font=("Segoe UI", 8),
                             bd=0, bg=c.get("card_bg", "#f0efe8"), fg="#5f5e5a",
                             pady=2, padx=6, cursor="hand2", command=cmd)

        _pbtn("▶▶ Última",   lambda: self._ir_pagina(self._total_paginas)).pack(side="right", padx=2)
        _pbtn("Siguiente ▶", lambda: self._ir_pagina(self._pagina + 1)).pack(side="right", padx=2)

        self._v_pag_entry = tk.StringVar()
        e_pag = ttk.Entry(pag, textvariable=self._v_pag_entry,
                          width=4, font=("Segoe UI", 8), justify="center")
        e_pag.pack(side="right", padx=2)
        e_pag.bind("<Return>", self._ir_pagina_directa)
        e_pag.bind("<FocusOut>", self._ir_pagina_directa)
        tk.Label(pag, text="Ir a:", font=("Segoe UI", 8),
                 bg=c.get("card_bg", "#f0efe8"), fg="#888780").pack(side="right", padx=(4, 0))

        _pbtn("◀ Anterior", lambda: self._ir_pagina(self._pagina - 1)).pack(side="right", padx=2)
        _pbtn("◀◀ Primera", lambda: self._ir_pagina(1)).pack(side="right", padx=2)

    def _ir_pagina_directa(self, event=None):
        try:
            n = int(self._v_pag_entry.get())
            self._ir_pagina(n)
        except (ValueError, TypeError):
            pass

    def _ir_pagina(self, n):
        if n < 1:
            n = 1
        if n > self._total_paginas:
            n = self._total_paginas
        if n != self._pagina:
            self._pagina = n
            self._cargar_resumen()

    def _build_detalle_vacio(self):
        for w in self._detalle_frame.winfo_children():
            w.destroy()
        c = self.app.colores
        tk.Label(self._detalle_frame,
                 text="📅\n\nSelecciona un mes\npara ver el detalle",
                 font=("Segoe UI", 11), bg="#ffffff", fg="#ccc9c0",
                 justify="center").pack(expand=True)

    def refresh(self):
        self._mes_sel = None
        self._pagina = 1
        self._cargar_resumen()
        self._build_detalle_vacio()

    def _actualizar_cards(self, datos):
        total_meses = len(datos)
        total_envios = sum(f.get("envios", 0) for f in datos)
        total_vendido = sum(f.get("total_vendido") or 0 for f in datos)
        total_pagado = sum(f.get("total_pagado") or 0 for f in datos)
        total_pendiente = sum(f.get("pendiente") or 0 for f in datos)

        self._cards["meses"].config(text=str(total_meses))
        self._cards["envios"].config(text=str(total_envios))
        self._cards["vendido"].config(text=f"${total_vendido:,.2f}")
        self._cards["pagado"].config(text=f"${total_pagado:,.2f}")
        self._cards["pendiente"].config(text=f"${total_pendiente:,.2f}")

    def _cargar_resumen(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        self._datos_cache = self.db.obtener_resumen_mensual()
        self._actualizar_cards(self._datos_cache)

        total_registros = len(self._datos_cache)
        if total_registros == 0:
            self._total_paginas = 1
            self._pag_label.config(text="Página 1 de 1")
            return

        self._total_paginas = max(1, (total_registros + self._por_pagina - 1) // self._por_pagina)
        if self._pagina > self._total_paginas:
            self._pagina = self._total_paginas

        inicio = (self._pagina - 1) * self._por_pagina
        fin = inicio + self._por_pagina
        pagina_datos = self._datos_cache[inicio:fin]

        for i, fila in enumerate(pagina_datos):
            mes_num = fila["mes"]
            mes_nombre = MESES_ES.get(mes_num, str(mes_num))
            anio = fila["anio"]
            envios = fila["envios"]
            vendido = fila["total_vendido"] or 0
            pagado = fila["total_pagado"] or 0
            pendiente = fila["pendiente"] or 0

            tag = "par" if i % 2 == 0 else "impar"
            tags_extra = (tag,)
            if pendiente > 0:
                tags_extra = ("con_deuda", tag)

            self.tree.insert("", "end", iid=f"{anio}-{mes_num}",
                             values=(mes_nombre, anio, envios,
                                     f"${vendido:,.2f}",
                                     f"${pagado:,.2f}",
                                     f"${pendiente:,.2f}"),
                             tags=tags_extra)

        self._pag_label.config(
            text=f"Página {self._pagina} de {self._total_paginas}  ({total_registros} meses)")

    def _on_seleccionar(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        parts = iid.split("-")
        anio = int(parts[0])
        mes = int(parts[1])
        self._mes_sel = (anio, mes)
        self._cargar_detalle(anio, mes)

    def _cargar_detalle(self, anio, mes):
        for w in self._detalle_frame.winfo_children():
            w.destroy()

        c = self.app.colores

        # Header del detalle
        hdr = tk.Frame(self._detalle_frame, bg=c.get("card_header", "#f0efe8"), pady=12, padx=16)
        hdr.pack(fill="x")
        mes_nombre = MESES_ES.get(mes, str(mes))
        tk.Label(hdr, text=f"📋  {mes_nombre} {anio}",
                 font=("Segoe UI", 13, "bold"),
                 bg=c.get("card_header", "#f0efe8"),
                 fg=c.get("card_header_fg", "#0f6e56")).pack(side="left")

        envios = self.db.obtener_envios_por_mes(anio, mes)

        if not envios:
            frame_vacio = tk.Frame(self._detalle_frame, bg="#ffffff")
            frame_vacio.pack(fill="both", expand=True)
            tk.Label(frame_vacio, text="📦\n\nNo hay envíos\nen este mes",
                     font=("Segoe UI", 11), bg="#ffffff", fg="#ccc9c0",
                     justify="center").pack(expand=True)
            return

        # Stats del mes
        stats = tk.Frame(self._detalle_frame, bg="#ffffff", padx=16, pady=8)
        stats.pack(fill="x")

        total_v = sum(e.get("total", 0) for e in envios)
        total_p = sum(e.get("abono", 0) for e in envios)
        pend = total_v - total_p

        for txt, val, fg in [
            (f"📦 {len(envios)} envíos", "", "#333"),
            (f"💰 Vendido: ${total_v:,.2f}", "", "#0c447c"),
            (f"✅ Pagado: ${total_p:,.2f}", "", "#085041"),
            (f"⏳ Pendiente: ${pend:,.2f}", "", "#791f1f" if pend > 0 else "#085041"),
        ]:
            tk.Label(stats, text=txt, font=("Segoe UI", 9),
                     bg="#ffffff", fg=fg).pack(side="left", padx=(0, 16))

        style = ttk.Style()
        style.configure("DetMes.Treeview",
                         font=("Segoe UI", 9), rowheight=26,
                         background="#ffffff", fieldbackground="#ffffff",
                         borderwidth=0)
        style.configure("DetMes.Treeview.Heading",
                         font=("Segoe UI", 8, "bold"),
                         background="#0f6e56", foreground="#ffffff",
                         relief="flat")

        tree_frame = tk.Frame(self._detalle_frame, bg="#ffffff")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("codigo", "fecha", "entrega", "recibe", "total", "abono", "estado")
        tree = ttk.Treeview(tree_frame, columns=cols,
                             show="headings", selectmode="browse",
                             style="DetMes.Treeview")
        hdrs_det = [
            ("codigo",  "Código",    85, "w"),
            ("fecha",   "Fecha",     75, "w"),
            ("entrega", "Entrega",  100, "w"),
            ("recibe",  "Recibe",   100, "w"),
            ("total",   "Total",     80, "e"),
            ("abono",   "Abono",     80, "e"),
            ("estado",  "Estado",    70, "center"),
        ]
        for col, txt, w, anc in hdrs_det:
            tree.heading(col, text=txt)
            tree.column(col, width=w, minwidth=35, anchor=anc)

        tree.tag_configure("Pagado",    foreground="#085041", background="#f6fdf9")
        tree.tag_configure("Abono",     foreground="#0c447c", background="#f4f8fd")
        tree.tag_configure("Pendiente", foreground="#7a4800", background="#fdf8f0")
        tree.tag_configure("Cancelado", foreground="#791f1f", background="#fdf3f3")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)

        for e in envios:
            fecha = (e.get("fecha", "") or "")[:10]
            estado = e.get("estado", "")
            tree.insert("", "end",
                        values=(e.get("codigo", ""), fecha,
                                e.get("ent_nombre", ""), e.get("rec_nombre", ""),
                                f"${e.get('total', 0):,.2f}",
                                f"${e.get('abono', 0):,.2f}",
                                estado),
                        tags=(estado,))

    def _exportar_excel(self):
        if not self._datos_cache:
            messagebox.showinfo("Sin datos", "No hay datos para exportar")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Error", "Instala openpyxl: pip install openpyxl")
            return

        from tkinter import filedialog
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"historial_mensual_{datetime.now().strftime('%Y%m')}.xlsx"
        )
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hist. Mensual"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="0F6E56", end_color="0F6E56", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["Mes", "Año", "Envíos", "Vendido", "Pagado", "Pendiente"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

        for i, fila in enumerate(self._datos_cache, 2):
            mes_num = fila["mes"]
            mes_nombre = MESES_ES.get(mes_num, str(mes_num))
            ws.cell(row=i, column=1, value=mes_nombre).border = thin_border
            ws.cell(row=i, column=2, value=fila["anio"]).border = thin_border
            ws.cell(row=i, column=3, value=fila["envios"]).border = thin_border
            ws.cell(row=i, column=4, value=fila.get("total_vendido") or 0).border = thin_border
            ws.cell(row=i, column=5, value=fila.get("total_pagado") or 0).border = thin_border
            ws.cell(row=i, column=6, value=fila.get("pendiente") or 0).border = thin_border

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        wb.save(ruta)
        messagebox.showinfo("Éxito", f"Exportado a:\n{ruta}")

    def _imprimir(self):
        from modules.imprimir import imprimir_historial
        if self._datos_cache:
            imprimir_historial(self._datos_cache)


# Necesario para _exportar_excel
from datetime import datetime
