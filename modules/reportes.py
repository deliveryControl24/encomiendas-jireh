"""
==============================================
  modules/reportes.py
  Dashboard de reportes con estadisticas,
  graficas de pastel y tabla por mes
==============================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from modules.imprimir import imprimir_reporte
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


class ReportesFrame(tk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent, bg="#f5f5f0")
        self.db  = db
        self.app = app
        self._build()

    def _build(self):
        header = tk.Frame(self, bg="#ffffff", pady=14)
        header.pack(fill="x")
        tk.Label(header, text="📊  Reportes",
                 font=("Segoe UI", 14, "bold"),
                 bg="#ffffff", fg="#0f6e56").pack(side="left", padx=20)

        btn_ref = tk.Button(header, text="↺ Actualizar",
                             font=("Segoe UI", 9), bd=0,
                             bg="#e1f5ee", fg="#0f6e56",
                             activebackground="#9fe1cb",
                             pady=5, padx=12, cursor="hand2",
                             command=self.refresh)
        btn_ref.pack(side="right", padx=20)

        btn_print = tk.Button(header, text="🖨 Imprimir reporte",
                              font=("Segoe UI", 9), bd=0,
                              bg="#e1f5ee", fg="#0f6e56",
                              activebackground="#9fe1cb",
                              pady=5, padx=12, cursor="hand2",
                              command=self._imprimir)
        btn_print.pack(side="right", padx=(0, 8))

        btn_excel = tk.Button(header, text="📊 Exportar Excel",
                              font=("Segoe UI", 9), bd=0,
                              bg="#e1f5ee", fg="#0f6e56",
                              activebackground="#9fe1cb",
                              pady=5, padx=12, cursor="hand2",
                              command=self._exportar_excel)
        btn_excel.pack(side="right", padx=(0, 8))

        canvas = tk.Canvas(self, bg="#f5f5f0", highlightthickness=0)
        vsb    = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=16, pady=10)

        self.inner = tk.Frame(canvas, bg="#f5f5f0")
        win = canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>",
                        lambda e: canvas.configure(
                            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(win, width=e.width))

        self.refresh()

    def refresh(self):
        for w in self.inner.winfo_children():
            w.destroy()

        resumen = self.db.resumen_general()
        estados = self.db.envios_por_estado()
        meses   = self.db.envios_por_mes()
        deudas_vencidas = self.db.obtener_deudas_vencidas(dias=7)
        clientes_deuda = self.db.obtener_clientes_con_deuda()
        stats = self.db.stats_resumen()
        envios = self.db.listar_envios()
        peso_total = sum((e.get("peso_total", 0) or 0) for e in envios)

        self._tarjetas(resumen, peso_total)
        self._alerta_deudas_vencidas(deudas_vencidas, clientes_deuda)
        self._grafica_pastel(estados)
        self._tabla_meses(meses)
        self._tabla_estados(estados)

    def _imprimir(self):
        try:
            ruta = imprimir_reporte(self.db)
            messagebox.showinfo("PDF generado ✓",
                                f"Archivo guardado en:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al imprimir",
                                 f"No se pudo generar el PDF:\n{e}")

    def _exportar_excel(self):
        try:
            ruta = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")],
                title="Guardar reporte Excel",
                initialfile=f"reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            if not ruta:
                return

            wb = openpyxl.Workbook()
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0F6E56", end_color="0F6E56", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Hoja 1: Resumen
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"
            
            # Título
            ws_resumen.merge_cells("A1:E1")
            ws_resumen["A1"] = "REPORTE DE ENCOMIENDAS"
            ws_resumen["A1"].font = Font(bold=True, size=14, color="0F6E56")
            ws_resumen["A1"].alignment = Alignment(horizontal="center")
            
            ws_resumen.merge_cells("A2:E2")
            ws_resumen["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws_resumen["A2"].alignment = Alignment(horizontal="center")
            
            # Datos del resumen
            stats = self.db.stats_resumen()
            resumen = self.db.resumen_general()
            
            envios = self.db.listar_envios()
            peso_total = sum((e.get("peso_total", 0) or 0) for e in envios)

            datos_resumen = [
                ("Total Envíos", stats["envios"]),
                ("Total Facturado", f"${stats['facturado']:,.2f}"),
                ("Total Cobrado", f"${stats['cobrado']:,.2f}"),
                ("Total Pendiente", f"${stats['pendiente']:,.2f}"),
                ("Total Costos", f"${stats['costos']:,.2f}"),
                ("Total en libras (lb)", f"{peso_total:,.2f}"),
            ]
            
            ws_resumen["A4"] = "Concepto"
            ws_resumen["B4"] = "Valor"
            ws_resumen["A4"].font = header_font
            ws_resumen["B4"].font = header_font
            ws_resumen["A4"].fill = header_fill
            ws_resumen["B4"].fill = header_fill
            ws_resumen["A4"].border = thin_border
            ws_resumen["B4"].border = thin_border
            
            for i, (concepto, valor) in enumerate(datos_resumen, start=5):
                ws_resumen[f"A{i}"] = concepto
                ws_resumen[f"B{i}"] = valor
                ws_resumen[f"A{i}"].border = thin_border
                ws_resumen[f"B{i}"].border = thin_border
            
            ws_resumen.column_dimensions["A"].width = 25
            ws_resumen.column_dimensions["B"].width = 20

            # Hoja 2: Detalle por Mes
            ws_meses = wb.create_sheet("Por Mes")
            meses = self.db.envios_por_mes()
            
            ws_meses["A1"] = "RESUMEN POR MES"
            ws_meses["A1"].font = Font(bold=True, size=12, color="0F6E56")
            
            headers_mes = ["Mes", "Envíos", "Monto Facturado"]
            for col, header in enumerate(headers_mes, start=1):
                cell = ws_meses.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for i, m in enumerate(meses, start=4):
                ws_meses.cell(row=i, column=1, value=m["mes"] or "Sin fecha").border = thin_border
                ws_meses.cell(row=i, column=2, value=m["cantidad"]).border = thin_border
                ws_meses.cell(row=i, column=3, value=f"${m['monto'] or 0:,.2f}").border = thin_border
            
            ws_meses.column_dimensions["A"].width = 15
            ws_meses.column_dimensions["B"].width = 12
            ws_meses.column_dimensions["C"].width = 20

            # Hoja 3: Detalle por Estado
            ws_estados = wb.create_sheet("Por Estado")
            estados = self.db.envios_por_estado()
            
            ws_estados["A1"] = "DETALLE POR ESTADO"
            ws_estados["A1"].font = Font(bold=True, size=12, color="0F6E56")
            
            headers_estado = ["Estado", "Cantidad", "Monto Total"]
            for col, header in enumerate(headers_estado, start=1):
                cell = ws_estados.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for i, e in enumerate(estados, start=4):
                ws_estados.cell(row=i, column=1, value=e["estado"]).border = thin_border
                ws_estados.cell(row=i, column=2, value=e["cantidad"]).border = thin_border
                ws_estados.cell(row=i, column=3, value=f"${e['monto']:,.2f}").border = thin_border
            
            ws_estados.column_dimensions["A"].width = 15
            ws_estados.column_dimensions["B"].width = 12
            ws_estados.column_dimensions["C"].width = 20

            # Hoja 4: Detalle de Envíos con Peso
            ws_envios = wb.create_sheet("Envíos con Peso")
            envios = self.db.listar_envios()
            
            ws_envios["A1"] = "DETALLE DE ENVÍOS CON PESO"
            ws_envios["A1"].font = Font(bold=True, size=12, color="0F6E56")
            
            headers_envios = ["Código", "Fecha", "Remitente", "Tel. Remitente", "Destinatario", "Tel. Destinatario", "Destino", "Peso (lb)", "Total", "Estado"]
            for col, header in enumerate(headers_envios, start=1):
                cell = ws_envios.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            total_peso = 0
            for i, env in enumerate(envios, start=4):
                peso = env.get("peso_total", 0) or 0
                total_peso += peso
                
                ws_envios.cell(row=i, column=1, value=env["codigo"]).border = thin_border
                ws_envios.cell(row=i, column=2, value=env["fecha"][:10]).border = thin_border
                ws_envios.cell(row=i, column=3, value=env["ent_nombre"]).border = thin_border
                ws_envios.cell(row=i, column=4, value=env.get("ent_tel", "")).border = thin_border
                ws_envios.cell(row=i, column=5, value=env["rec_nombre"]).border = thin_border
                ws_envios.cell(row=i, column=6, value=env.get("rec_tel", "")).border = thin_border
                ws_envios.cell(row=i, column=7, value=env["destino_usa"]).border = thin_border
                ws_envios.cell(row=i, column=8, value=f"{peso:.2f}").border = thin_border
                ws_envios.cell(row=i, column=9, value=f"${env['total']:,.2f}").border = thin_border
                ws_envios.cell(row=i, column=10, value=env["estado"]).border = thin_border
            
            # Fila de totales
            total_row = len(envios) + 4
            ws_envios.cell(row=total_row, column=7, value="TOTAL PESO:").font = Font(bold=True)
            ws_envios.cell(row=total_row, column=8, value=f"{total_peso:.2f} lb").font = Font(bold=True)
            ws_envios.cell(row=total_row, column=7).border = thin_border
            ws_envios.cell(row=total_row, column=8).border = thin_border
            
            for col in range(1, 11):
                ws_envios.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

            wb.save(ruta)
            messagebox.showinfo("Excel generado ✓",
                                f"Archivo guardado en:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al exportar",
                                 f"No se pudo generar el Excel:\n{e}")

    # ─── Tarjetas de resumen ────────────────────────────────────────────────

    def _tarjetas(self, r, peso_total=0):
        frame = tk.Frame(self.inner, bg="#f5f5f0")
        frame.pack(fill="x", pady=(0, 12))

        datos = [
            ("Total envíos",    str(int(r["total_envios"])),  "#e1f5ee", "#085041"),
            ("Facturado",       f"${r['total_facturado']:,.0f}", "#e6f1fb", "#0c447c"),
            ("Cobrado",         f"${r['total_cobrado']:,.0f}",   "#eaf3de", "#27500a"),
            ("Pendiente",       f"${r['total_pendiente']:,.0f}", "#faeeda", "#633806"),
            ("Total en libras (lb)", f"{peso_total:,.2f}", "#f3e8ff", "#6b21a8"),
        ]

        for i, (lbl, val, bg, fg) in enumerate(datos):
            card = tk.Frame(frame, bg=bg,
                            highlightthickness=1,
                            highlightbackground="#d3d1c7")
            card.grid(row=0, column=i, padx=(0, 8) if i < 4 else 0,
                      sticky="nsew")
            frame.columnconfigure(i, weight=1)

            tk.Label(card, text=lbl, font=("Segoe UI", 9),
                     bg=bg, fg=fg).pack(padx=14, pady=(12, 2), anchor="w")
            tk.Label(card, text=val, font=("Segoe UI", 16, "bold"),
                     bg=bg, fg=fg).pack(padx=14, pady=(0, 12), anchor="w")

    # ─── Alerta de deudas vencidas ──────────────────────────────────────────

    def _alerta_deudas_vencidas(self, deudas, clientes_deuda):
        if not clientes_deuda:
            return

        card = tk.Frame(self.inner, bg="#fcebeb",
                        highlightthickness=1,
                        highlightbackground="#e0a8a8", relief="flat")
        card.pack(fill="x", pady=(0, 12))

        header_frame = tk.Frame(card, bg="#f5d4d4")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="⚠️  ALERTA: Clientes con deuda pendiente",
                 font=("Segoe UI", 9, "bold"),
                 bg="#f5d4d4", fg="#791f1f",
                 pady=8, padx=14, anchor="w").pack(fill="x")

        body = tk.Frame(card, bg="#fcebeb", padx=14, pady=10)
        body.pack(fill="x")

        total_deudores = len(clientes_deuda)
        monto_total = sum(c["total_deuda"] for c in clientes_deuda)

        resumen_txt = f"Hay {total_deudores} cliente(s) con deuda total de $ {monto_total:,.2f}"
        tk.Label(body, text=resumen_txt,
                 font=("Segoe UI", 10, "bold"),
                 bg="#fcebeb", fg="#791f1f").pack(anchor="w", pady=(0, 10))

        tree = ttk.Treeview(body,
                             columns=("cliente", "deudas", "total", "desde"),
                             show="headings", height=min(len(clientes_deuda), 5))
        tree.heading("cliente", text="Cliente")
        tree.heading("deudas",  text="Deudas")
        tree.heading("total",   text="Monto")
        tree.heading("desde",   text="Desde")
        tree.column("cliente", width=150, anchor="w")
        tree.column("deudas",  width=60,  anchor="center")
        tree.column("total",   width=100, anchor="e")
        tree.column("desde",   width=100, anchor="center")

        for c in clientes_deuda:
            tree.insert("", "end", values=(
                c["ent_nombre"],
                c["cantidad_deudas"],
                f"$ {c['total_deuda']:,.2f}",
                c["deuda_desde"][:10],
            ), tags=("deuda",))

        tree.pack(fill="x", pady=8)
        tree.tag_configure("deuda", background="#faeeda", foreground="#633806")

    # ─── Gráfico de pastel ─────────────────────────────────────────────────

    def _grafica_pastel(self, estados):
        card = tk.Frame(self.inner, bg="#ffffff",
                        highlightthickness=1,
                        highlightbackground="#e0e0d8")
        card.pack(fill="x", pady=(0, 12))

        tk.Label(card, text="DISTRIBUCIÓN POR ESTADO",
                 font=("Segoe UI", 8, "bold"),
                 bg="#f1efe8", fg="#5f5e5a",
                 pady=8, padx=14, anchor="w").pack(fill="x")

        cvs = tk.Canvas(card, bg="#ffffff", height=220,
                         highlightthickness=0)
        cvs.pack(fill="x", padx=20, pady=14)

        colores = {
            "Pagado":    "#1d9e75",
            "Abono":     "#378add",
            "Pendiente": "#ef9f27",
            "Cancelado": "#e24b4a",
        }

        total = sum(e["cantidad"] for e in estados) or 1

        cvs.update_idletasks()
        ancho = cvs.winfo_width() or 600
        cx, cy = ancho // 2 - 40, 110
        radio = 75

        # Ordenar estados para consistencia
        orden = ["Pagado", "Abono", "Pendiente", "Cancelado"]
        estados_ord = [e for e in orden if any(x["estado"] == e for x in estados)]
        estados_ord = [next(x for x in estados if x["estado"] == e) for e in estados_ord if any(x["estado"] == e for x in estados)]
        # Add any estados not in the predefined order
        extra = [e for e in estados if e["estado"] not in orden]
        estados_ord.extend(extra)

        angulo_inicio = 90
        for e in estados_ord:
            proporcion = e["cantidad"] / total
            angulo_barrido = proporcion * 360
            color = colores.get(e["estado"], "#888780")
            cvs.create_arc(
                cx - radio, cy - radio,
                cx + radio, cy + radio,
                start=angulo_inicio,
                extent=angulo_barrido,
                fill=color, outline="#ffffff", width=2
            )
            angulo_medio = angulo_inicio + angulo_barrido / 2
            import math
            ang_rad = math.radians(angulo_medio)
            etq_r = radio * 0.65
            lx = cx + etq_r * math.cos(ang_rad)
            ly = cy - etq_r * math.sin(ang_rad)
            if proporcion >= 0.05:
                cvs.create_text(lx, ly, text=f"{proporcion*100:.0f}%",
                                font=("Segoe UI", 9, "bold"),
                                fill="#ffffff", anchor="center")
            angulo_inicio += angulo_barrido

        # Leyenda
        lx_start = cx + radio + 30
        ly_start = cy - radio
        for i, e in enumerate(estados_ord):
            y = ly_start + i * 22
            color = colores.get(e["estado"], "#888780")
            cvs.create_rectangle(lx_start, y, lx_start + 14, y + 14,
                                  fill=color, outline="")
            cvs.create_text(lx_start + 20, y + 7, anchor="w",
                             text=f"{e['estado']} ({e['cantidad']})",
                             font=("Segoe UI", 9), fill="#444441")

    # ─── Tabla por mes ───────────────────────────────────────────────────────

    def _tabla_meses(self, meses):
        card = tk.Frame(self.inner, bg="#ffffff",
                        highlightthickness=1,
                        highlightbackground="#e0e0d8")
        card.pack(fill="x", pady=(0, 12))

        tk.Label(card, text="RESUMEN POR MES",
                 font=("Segoe UI", 8, "bold"),
                 bg="#f1efe8", fg="#5f5e5a",
                 pady=8, padx=14, anchor="w").pack(fill="x")

        if not meses:
            tk.Label(card, text="Sin datos aún.",
                     font=("Segoe UI", 9), bg="#ffffff",
                     fg="#888780").pack(pady=14)
            return

        tree = ttk.Treeview(card,
                             columns=("mes", "envios", "monto"),
                             show="headings", height=min(len(meses), 7))
        tree.heading("mes",    text="Mes")
        tree.heading("envios", text="Envíos")
        tree.heading("monto",  text="Monto facturado")
        tree.column("mes",    width=120, anchor="w")
        tree.column("envios", width=80,  anchor="center")
        tree.column("monto",  width=180, anchor="e")

        for m in meses:
            mes_label = m["mes"] if m["mes"] else "Sin fecha"
            monto = m["monto"] or 0
            tree.insert("", "end", values=(
                mes_label,
                m["cantidad"],
                f"${monto:,.2f}",
            ))

        tree.pack(fill="x", padx=14, pady=(8, 14))

    # ─── Tabla por estado ────────────────────────────────────────────────────

    def _tabla_estados(self, estados):
        card = tk.Frame(self.inner, bg="#ffffff",
                        highlightthickness=1,
                        highlightbackground="#e0e0d8")
        card.pack(fill="x", pady=(0, 16))

        tk.Label(card, text="DETALLE POR ESTADO",
                 font=("Segoe UI", 8, "bold"),
                 bg="#f1efe8", fg="#5f5e5a",
                 pady=8, padx=14, anchor="w").pack(fill="x")

        if not estados:
            tk.Label(card, text="Sin datos aún.",
                     font=("Segoe UI", 9), bg="#ffffff",
                     fg="#888780").pack(pady=14)
            return

        tree = ttk.Treeview(card,
                             columns=("estado", "cantidad", "monto"),
                             show="headings", height=len(estados))
        tree.heading("estado",   text="Estado")
        tree.heading("cantidad", text="Cantidad")
        tree.heading("monto",    text="Total")
        tree.column("estado",   width=120, anchor="w")
        tree.column("cantidad", width=80,  anchor="center")
        tree.column("monto",    width=180, anchor="e")

        colores_fg = {
            "Pagado":    "#085041",
            "Abono":     "#0c447c",
            "Pendiente": "#633806",
            "Cancelado": "#791f1f",
        }
        for e in estados:
            iid = tree.insert("", "end", values=(
                e["estado"],
                e["cantidad"],
                f"${e['monto']:,.2f}",
            ))

        tree.pack(fill="x", padx=14, pady=(8, 14))
