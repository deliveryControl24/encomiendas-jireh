"""
==============================================
  modules/historial_clientes.py
  Directorio de clientes con historial de envios
  Cards resumen, exportar Excel, mejor detalle
==============================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from modules.config import PAGINA_TAMANO
import os
import sys
from datetime import datetime

try:
    from modules.config import get_base_dir
except ImportError:
    def get_base_dir():
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


class HistorialClientesFrame(tk.Frame):
    def __init__(self, parent, db, app):
        super().__init__(parent, bg="#f5f5f0")
        self.db = db
        self.app = app
        self._pagina = 1
        self._por_pagina = PAGINA_TAMANO
        self._total_paginas = 1
        self._cliente_sel = None
        self._orden_col = "nombre"
        self._orden_reverso = False
        self._datos_cache = []
        self._build()

    def _build(self):
        c = self.app.colores

        # ── Header ───────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=c.get("card_bg", "#ffffff"), pady=14, padx=20)
        hdr.pack(fill="x")
        tk.Label(hdr, text="👤  Historial de clientes",
                 font=("Segoe UI", 14, "bold"),
                 bg=c.get("card_bg", "#ffffff"),
                 fg=c.get("accent", "#0f6e56")).pack(side="left")

        tk.Button(hdr, text="📥  Exportar Excel",
                  font=("Segoe UI", 9, "bold"),
                  bg=c.get("accent", "#0f6e56"), fg="#ffffff",
                  bd=0, padx=14, pady=6, cursor="hand2",
                  command=self._exportar_excel).pack(side="right", padx=(8, 0))

        # ── Cards resumen ────────────────────────────────────────────────────
        self.cards_frame = tk.Frame(self, bg=c.get("bg", "#f5f5f0"))
        self.cards_frame.pack(fill="x", padx=14, pady=(10, 0))

        self._cards = {}
        labels_cards = [
            ("clientes",     "👤  Clientes totales",  "0",     "#0f6e56", "#e1f5ee"),
            ("envios",       "📦  Envíos totales",    "0",     "#0c447c", "#e6f1fb"),
            ("vendido",      "💰  Total vendido",      "$0.00", "#7a4800", "#faeeda"),
            ("pagado",       "✅  Total pagado",       "$0.00", "#085041", "#f6fdf9"),
            ("top_cliente",  "⭐  Top cliente",        "—",     "#8b5a00", "#fdf8ed"),
        ]
        for key, titulo, valor, fg_color, bg_color in labels_cards:
            card = tk.Frame(self.cards_frame, bg=bg_color,
                            highlightthickness=1, highlightbackground=fg_color + "33",
                            padx=14, pady=10)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8))
            tk.Label(card, text=titulo, font=("Segoe UI", 9),
                     bg=bg_color, fg=fg_color).pack(anchor="w")
            lbl_valor = tk.Label(card, text=valor,
                                 font=("Segoe UI", 13, "bold"),
                                 bg=bg_color, fg=fg_color,
                                 wraplength=120, justify="left")
            lbl_valor.pack(anchor="w", pady=(2, 0))
            self._cards[key] = lbl_valor

        # ── Barra de búsqueda ────────────────────────────────────────────────
        bar = tk.Frame(self, bg="#f0efe8", padx=14, pady=6)
        bar.pack(fill="x")

        tk.Label(bar, text="🔍  Buscar:", font=("Segoe UI", 9),
                 bg="#f0efe8", fg="#888780").pack(side="left", padx=(0, 4))
        self.v_buscar = tk.StringVar()
        self.v_buscar.trace_add("write", lambda *a: self._ir_pagina(1))
        ttk.Entry(bar, textvariable=self.v_buscar,
                  width=24, font=("Segoe UI", 10)).pack(side="left", padx=(0, 12))

        tk.Label(bar, text="📞  Teléfono:", font=("Segoe UI", 9),
                 bg="#f0efe8", fg="#888780").pack(side="left", padx=(0, 4))
        self.v_tel = tk.StringVar()
        self.v_tel.trace_add("write", lambda *a: self._ir_pagina(1))
        ttk.Entry(bar, textvariable=self.v_tel,
                  width=14, font=("Segoe UI", 10)).pack(side="left")

        tk.Button(bar, text="↺ Limpiar",
                  font=("Segoe UI", 9), bd=0,
                  bg="#e1f5ee", fg="#0f6e56", padx=10, pady=4,
                  cursor="hand2", command=self._limpiar_busqueda).pack(side="right")

        # ── Pane principal ───────────────────────────────────────────────────
        paned = tk.PanedWindow(self, orient="horizontal",
                                bg="#d8d6cf", sashwidth=5, sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=14, pady=(6, 12))

        left = tk.Frame(paned, bg="#ffffff",
                        highlightthickness=1, highlightbackground="#dddbd4")
        paned.add(left, minsize=420)

        self._build_tree(left)
        self._build_paginacion(left)

        self._detalle_frame = tk.Frame(paned, bg="#ffffff",
                                        highlightthickness=1, highlightbackground="#dddbd4")
        paned.add(self._detalle_frame, minsize=320)
        self._build_detalle_vacio()

        self.refresh()

    def _limpiar_busqueda(self):
        self.v_buscar.set("")
        self.v_tel.set("")

    def _build_tree(self, parent):
        style = ttk.Style()
        style.configure("Cli.Treeview",
                         font=("Segoe UI", 9), rowheight=30,
                         background="#ffffff", fieldbackground="#ffffff",
                         borderwidth=0)
        style.configure("Cli.Treeview.Heading",
                         font=("Segoe UI", 9, "bold"),
                         background="#0f6e56", foreground="#ffffff",
                         relief="flat")
        style.map("Cli.Treeview",
                  background=[("selected", "#d4f5eb")],
                  foreground=[("selected", "#085041")])

        cols = ("nombre", "telefono", "envios", "total_gastado", "ultimo_envio")
        self.tree = ttk.Treeview(parent, columns=cols,
                                  show="headings", selectmode="browse",
                                  style="Cli.Treeview")
        hdrs = [
            ("nombre",       "Nombre",       130, "w"),
            ("telefono",     "Teléfono",      95, "w"),
            ("envios",       "Envíos",        55, "center"),
            ("total_gastado","Total gastado", 90, "e"),
            ("ultimo_envio", "Último envío",  85, "center"),
        ]
        for col, txt, w, anc in hdrs:
            self.tree.heading(col, text=txt,
                              command=lambda c=col: self._ordenar(c))
            self.tree.column(col, width=w, minwidth=40, anchor=anc)

        self.tree.tag_configure("par",   background="#f9f9f6")
        self.tree.tag_configure("impar", background="#ffffff")
        self.tree.tag_configure("top",   foreground="#8b5a00", background="#fdf8ed")

        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_seleccionar)

    def _ordenar(self, col):
        if self._orden_col == col:
            self._orden_reverso = not self._orden_reverso
        else:
            self._orden_col = col
            self._orden_reverso = False
        self._cargar_clientes()

    def _build_paginacion(self, parent):
        c = self.app.colores
        pag = tk.Frame(parent, bg=c.get("card_bg", "#f0efe8"), pady=5)
        pag.pack(fill="x")
        self._pag_label = tk.Label(pag, text="", font=("Segoe UI", 8),
                                    bg=c.get("card_bg", "#f0efe8"), fg="#888780")
        self._pag_label.pack(side="left", padx=10)

        def _pbtn(txt, cmd):
            return tk.Button(pag, text=txt, font=("Segoe UI", 8),
                             bd=0, bg=c.get("card_bg", "#f0efe8"), fg="#5f5e5a",
                             pady=2, padx=6, cursor="hand2", command=cmd)

        _pbtn("▶▶ Última",   lambda: self._ir_pagina(self._total_paginas)).pack(side="right", padx=2)
        _pbtn("Siguiente ▶", lambda: self._ir_pagina(self._pagina + 1)).pack(side="right", padx=2)

        self._v_pag_entry = tk.StringVar()
        e_pag = ttk.Entry(pag, textvariable=self._v_pag_entry,
                          width=4, font=("Segoe UI", 8), justify="center")
        e_pag.pack(side="right", padx=2)
        e_pag.bind("<Return>", self._ir_pagina_directa)
        e_pag.bind("<FocusOut>", self._ir_pagina_directa)
        tk.Label(pag, text="Ir a:", font=("Segoe UI", 8),
                 bg=c.get("card_bg", "#f0efe8"), fg="#888780").pack(side="right", padx=(4, 0))

        _pbtn("◀ Anterior", lambda: self._ir_pagina(self._pagina - 1)).pack(side="right", padx=2)
        _pbtn("◀◀ Primera", lambda: self._ir_pagina(1)).pack(side="right", padx=2)

    def _ir_pagina_directa(self, event=None):
        try:
            n = int(self._v_pag_entry.get())
            self._ir_pagina(n)
        except (ValueError, TypeError):
            pass

    def _ir_pagina(self, n):
        if n < 1:
            n = 1
        if n > self._total_paginas:
            n = self._total_paginas
        if n != self._pagina:
            self._pagina = n
            self._cargar_clientes()

    def _build_detalle_vacio(self):
        for w in self._detalle_frame.winfo_children():
            w.destroy()
        tk.Label(self._detalle_frame,
                 text="👤\n\nSelecciona un cliente\npara ver su historial",
                 font=("Segoe UI", 11), bg="#ffffff", fg="#ccc9c0",
                 justify="center").pack(expand=True)

    def refresh(self):
        self._cliente_sel = None
        self._pagina = 1
        self._cargar_clientes()
        self._build_detalle_vacio()

    def _actualizar_cards(self, datos):
        total_clientes = len(datos)
        total_envios = sum(f.get("envios", 0) for f in datos)
        total_vendido = sum(f.get("total_gastado") or 0 for f in datos)

        # Calcular top cliente
        top_cliente = "—"
        if datos:
            top = max(datos, key=lambda x: x.get("total_gastado") or 0)
            if top.get("total_gastado"):
                nombre = top.get("nombre", "")
                if len(nombre) > 12:
                    nombre = nombre[:12] + "…"
                top_cliente = f"{nombre}"

        self._cards["clientes"].config(text=str(total_clientes))
        self._cards["envios"].config(text=str(total_envios))
        self._cards["vendido"].config(text=f"${total_vendido:,.2f}")
        self._cards["top_cliente"].config(text=top_cliente)

    def _cargar_clientes(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        buscar = self.v_buscar.get().strip()
        tel = self.v_tel.get().strip()
        self._datos_cache = self.db.obtener_clientes(buscar=buscar, telefono=tel)
        self._actualizar_cards(self._datos_cache)

        if not self._datos_cache:
            self._total_paginas = 1
            self._pag_label.config(text="Página 1 de 1")
            return

        # Ordenar
        clave = self._orden_col
        reverse = self._orden_reverso
        if clave == "nombre":
            self._datos_cache.sort(key=lambda x: (x.get("nombre") or "").lower(), reverse=reverse)
        elif clave == "telefono":
            self._datos_cache.sort(key=lambda x: (x.get("telefono") or "").lower(), reverse=reverse)
        elif clave == "envios":
            self._datos_cache.sort(key=lambda x: x.get("envios", 0), reverse=reverse)
        elif clave == "total_gastado":
            self._datos_cache.sort(key=lambda x: x.get("total_gastado", 0), reverse=reverse)
        elif clave == "ultimo_envio":
            self._datos_cache.sort(key=lambda x: x.get("ultimo_envio") or "", reverse=reverse)

        total_registros = len(self._datos_cache)
        self._total_paginas = max(1, (total_registros + self._por_pagina - 1) // self._por_pagina)
        if self._pagina > self._total_paginas:
            self._pagina = self._total_paginas

        inicio = (self._pagina - 1) * self._por_pagina
        fin = inicio + self._por_pagina
        pagina_datos = self._datos_cache[inicio:fin]

        # Encontrar max envíos para marcar top
        max_envios = max((f.get("envios", 0) for f in self._datos_cache), default=0)

        for i, fila in enumerate(pagina_datos):
            nombre = fila.get("nombre", "")
            telefono = fila.get("telefono", "") or "—"
            envios = fila.get("envios", 0)
            total = fila.get("total_gastado", 0) or 0
            ultimo = (fila.get("ultimo_envio") or "")[:10]

            tag = "par" if i % 2 == 0 else "impar"
            if envios == max_envios and max_envios > 0:
                tag = "top"

            self.tree.insert("", "end", iid=nombre,
                             values=(nombre, telefono, envios,
                                     f"${total:,.2f}", ultimo),
                             tags=(tag,))

        self._pag_label.config(
            text=f"Página {self._pagina} de {self._total_paginas}  ({total_registros} clientes)")

    def _on_seleccionar(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        nombre = sel[0]
        self._cliente_sel = nombre
        self._cargar_detalle(nombre)

    def _cargar_detalle(self, nombre):
        for w in self._detalle_frame.winfo_children():
            w.destroy()

        c = self.app.colores

        envios = self.db.obtener_envios_por_cliente(nombre)

        # Header del detalle
        hdr = tk.Frame(self._detalle_frame, bg=c.get("card_header", "#f0efe8"), pady=12, padx=16)
        hdr.pack(fill="x")

        tel = ""
        if envios:
            tel = envios[0].get("ent_tel", "") or envios[0].get("rec_tel", "")
        tel_text = f"  📞 {tel}" if tel else ""

        tk.Label(hdr, text=f"👤  {nombre}{tel_text}",
                 font=("Segoe UI", 13, "bold"),
                 bg=c.get("card_header", "#f0efe8"),
                 fg=c.get("card_header_fg", "#0f6e56")).pack(side="left")

        if not envios:
            frame_vacio = tk.Frame(self._detalle_frame, bg="#ffffff")
            frame_vacio.pack(fill="both", expand=True)
            tk.Label(frame_vacio, text="📦\n\nNo hay envíos\npara este cliente",
                     font=("Segoe UI", 11), bg="#ffffff", fg="#ccc9c0",
                     justify="center").pack(expand=True)
            return

        # Stats del cliente
        stats = tk.Frame(self._detalle_frame, bg="#ffffff", padx=16, pady=8)
        stats.pack(fill="x")

        total_v = sum(e.get("total", 0) for e in envios)
        total_p = sum(e.get("abono", 0) for e in envios)
        pend = total_v - total_p

        for txt, fg in [
            (f"📦 {len(envios)} envíos", "#333"),
            (f"💰 Total: ${total_v:,.2f}", "#0c447c"),
            (f"✅ Pagado: ${total_p:,.2f}", "#085041"),
            (f"⏳ Pendiente: ${pend:,.2f}", "#791f1f" if pend > 0 else "#085041"),
        ]:
            tk.Label(stats, text=txt, font=("Segoe UI", 9),
                     bg="#ffffff", fg=fg).pack(side="left", padx=(0, 16))

        style = ttk.Style()
        style.configure("CliDet.Treeview",
                         font=("Segoe UI", 9), rowheight=26,
                         background="#ffffff", fieldbackground="#ffffff",
                         borderwidth=0)
        style.configure("CliDet.Treeview.Heading",
                         font=("Segoe UI", 8, "bold"),
                         background="#0f6e56", foreground="#ffffff",
                         relief="flat")

        tree_frame = tk.Frame(self._detalle_frame, bg="#ffffff")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ("codigo", "fecha", "recibe", "total", "abono", "estado")
        tree = ttk.Treeview(tree_frame, columns=cols,
                             show="headings", selectmode="browse",
                             style="CliDet.Treeview")
        hdrs_det = [
            ("codigo",  "Código",   85, "w"),
            ("fecha",   "Fecha",    75, "w"),
            ("recibe",  "Recibe",  100, "w"),
            ("total",   "Total",    80, "e"),
            ("abono",   "Abono",    80, "e"),
            ("estado",  "Estado",   70, "center"),
        ]
        for col, txt, w, anc in hdrs_det:
            tree.heading(col, text=txt)
            tree.column(col, width=w, minwidth=35, anchor=anc)

        tree.tag_configure("Pagado",    foreground="#085041", background="#f6fdf9")
        tree.tag_configure("Abono",     foreground="#0c447c", background="#f4f8fd")
        tree.tag_configure("Pendiente", foreground="#7a4800", background="#fdf8f0")
        tree.tag_configure("Cancelado", foreground="#791f1f", background="#fdf3f3")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)

        for e in envios:
            fecha = (e.get("fecha", "") or "")[:10]
            estado = e.get("estado", "")
            recibe = e.get("rec_nombre", "") if nombre != e.get("rec_nombre", "") else e.get("ent_nombre", "")
            tree.insert("", "end",
                        values=(e.get("codigo", ""), fecha, recibe,
                                f"${e.get('total', 0):,.2f}",
                                f"${e.get('abono', 0):,.2f}",
                                estado),
                        tags=(estado,))

        # Pie con totales
        pie = tk.Frame(self._detalle_frame, bg=c.get("card_header", "#f0efe8"), pady=8, padx=16)
        pie.pack(fill="x")
        tk.Label(pie, text=f"Envíos: {len(envios)}  |  "
                           f"Total: ${total_v:,.2f}  |  "
                           f"Pagado: ${total_p:,.2f}  |  "
                           f"Pendiente: ${pend:,.2f}",
                 font=("Segoe UI", 9),
                 bg=c.get("card_header", "#f0efe8"),
                 fg=c.get("card_header_fg", "#0f6e56")).pack(side="left")

    def _exportar_excel(self):
        if not self._datos_cache:
            messagebox.showinfo("Sin datos", "No hay datos para exportar")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Error", "Instala openpyxl: pip install openpyxl")
            return

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"clientes_{datetime.now().strftime('%Y%m')}.xlsx"
        )
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="0F6E56", end_color="0F6E56", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["Nombre", "Teléfono", "Envíos", "Total gastado", "Último envío"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

        for i, fila in enumerate(self._datos_cache, 2):
            ws.cell(row=i, column=1, value=fila.get("nombre", "")).border = thin_border
            ws.cell(row=i, column=2, value=fila.get("telefono", "") or "").border = thin_border
            ws.cell(row=i, column=3, value=fila.get("envios", 0)).border = thin_border
            ws.cell(row=i, column=4, value=fila.get("total_gastado") or 0).border = thin_border
            ws.cell(row=i, column=5, value=(fila.get("ultimo_envio") or "")[:10]).border = thin_border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

        wb.save(ruta)
        messagebox.showinfo("Éxito", f"Exportado a:\n{ruta}")
